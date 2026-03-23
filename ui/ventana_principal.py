import pandas as pd
import requests

from PySide6.QtWidgets import (
    QMainWindow, QWidget, QHBoxLayout, QVBoxLayout,
    QPushButton, QLabel, QFileDialog, QListWidget,
    QTableWidget, QTableWidgetItem, QLineEdit,
    QComboBox, QProgressBar, QStackedWidget,
    QGraphicsOpacityEffect, QListWidgetItem, QFrame
)

from cliente_api.api_client import conciliar_local

from PySide6.QtCore import Qt, QPropertyAnimation, QEasingCurve

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

from ui.descargas import DescargaChrome


class VentanaPrincipal(QMainWindow):

    def __init__(self):
        super().__init__()

        self.archivos = []

        self.central = QWidget()
        self.setCentralWidget(self.central)

        self.layout = QHBoxLayout()
        self.central.setLayout(self.layout)

        # SIDEBAR
        sidebar_widget = QWidget()
        sidebar_widget.setObjectName("sidebar")
        sidebar_widget.setFixedWidth(200)
        
        self.sidebar = QVBoxLayout(sidebar_widget)
        
        # BOTON MENU
        self.btn_menu = QPushButton("☰ Menú")
        self.btn_menu.clicked.connect(self.toggle_menu)
        self.sidebar.addWidget(self.btn_menu)
        
        # SUBMENU
        self.submenu = QWidget()
        submenu_layout = QVBoxLayout(self.submenu)
        
        self.btn_inicio = QPushButton("🏠 Inicio")
        self.btn_conciliador = QPushButton("📦 Conciliador")
        self.btn_descargas = QPushButton("⬇ Descargas")
        
        submenu_layout.addWidget(self.btn_inicio)
        submenu_layout.addWidget(self.btn_conciliador)
        submenu_layout.addWidget(self.btn_descargas)
        
        self.submenu.setMaximumHeight(0)
        self.sidebar.addWidget(self.submenu)
        self.sidebar.addStretch()
        self.layout.addWidget(sidebar_widget)

        # AREA CONTENIDO
        self.contenido = QStackedWidget()
        self.layout.addWidget(self.contenido)

        # PAGINA INICIO
        self.pagina_inicio = QWidget()
        layout_inicio = QVBoxLayout()
        
        label = QLabel("Bienvenido a Logistic Soft ")
        label.setAlignment(Qt.AlignCenter)
        
        layout_inicio.addWidget(label)
        self.pagina_inicio.setLayout(layout_inicio)
        self.contenido.addWidget(self.pagina_inicio)

        # TITULO
        self.titulo = QLabel("📦 Conciliador de Inventarios")
        self.titulo.setStyleSheet("""
        font-size:22px;
        font-weight:bold;
        color:#38bdf8;
        """)
        
        # ---------- PAGINA CONCILIADOR ----------
        self.pagina_conciliador = QWidget()
        layout_conciliador = QVBoxLayout()

        botones = QHBoxLayout()

        self.boton_upload = QPushButton("Upload")
        self.boton_upload.clicked.connect(self.subir_archivo)

        self.boton_conciliar = QPushButton("Conciliar")
        def ejecutar_conciliacion(self):
            if not self.archivos:
                print("No hay archivos")
                return
            self.boton_conciliar.setEnabled(False)
            self.boton_upload.setEnabled(False)
            
            try:
                print("Iniciando conciliación LOCAL...")
                
                resultado = conciliar_local(self.archivos)
                
                if resultado is not None:
                    print("Conciliación terminada")
                    
                    enviar_resultado(resultado)
                    
                    self.mostrar_tabla(resultado)
                else:
                    print("Sin resultados")
            except Exception as e:
                print("Error:", str(e))
                
                self.boton_conciliar.setEnabled(True)
                self.boton_upload.setEnabled(True)

        botones.addWidget(self.boton_upload)
        botones.addWidget(self.boton_conciliar)

        layout_conciliador.addLayout(botones)

        self.lista_archivos = QListWidget()
        layout_conciliador.addWidget(self.lista_archivos)

        self.progress = QProgressBar()
        layout_conciliador.addWidget(self.progress)

        self.buscador = QLineEdit()
        self.buscador.setPlaceholderText("Buscar en la tabla...")
        self.buscador.textChanged.connect(self.buscar_tabla)
        layout_conciliador.addWidget(self.buscador)

        self.filtro_columna = QComboBox()
        self.filtro_columna.addItem("Todas")
        layout_conciliador.addWidget(self.filtro_columna)

        self.tabla = QTableWidget()
        layout_conciliador.addWidget(self.tabla)

        self.pagina_conciliador.setLayout(layout_conciliador)
        self.contenido.addWidget(self.pagina_conciliador)

        # ---------- PAGINA DESCARGAS ----------
        self.pagina_descargas = QWidget()
        layout_descargas = QVBoxLayout()
        
        titulo = QLabel("Descargas")
        titulo.setStyleSheet("font-size:18px;font-weight:bold")
        
        layout_descargas.addWidget(titulo)
        
        self.lista_descargas = QListWidget()
        layout_descargas.addWidget(self.lista_descargas)
        
        self.boton_exportar = QPushButton("⬇ Exportar conciliación")
        self.boton_exportar.clicked.connect(self.exportar_excel)
        
        layout_descargas.addWidget(self.boton_exportar)
        
        self.pagina_descargas.setLayout(layout_descargas)
        self.contenido.addWidget(self.pagina_descargas)

        # CONEXIONES MENU
        self.btn_inicio.clicked.connect(
            lambda: self.cambiar_pagina(self.pagina_inicio)
        )
        self.btn_conciliador.clicked.connect(
            lambda: self.cambiar_pagina(self.pagina_conciliador)
        )
        self.btn_descargas.clicked.connect(
            lambda: self.cambiar_pagina(self.pagina_descargas)
        )

        # CONFIGURACION
        self.setWindowTitle("Logistic Soft")
        self.resize(900,600)


        self.buscador.hide()
        self.filtro_columna.hide()
        self.tabla.hide()

        self.detalle_componentes = None
        self.detalle_restricciones = None

    def enviar_resultado(resultado_df):
        URL_GUARDAR = "http://TU_IP/api/guardar/"

        try:
            data = resultado_df.to_dict(orient="records")
            response = requests.post(URL_GUARDAR, json=data)
            
            if response.status_code == 200:
                print("Guardado OK")
            else:
                print("Error servidor:", response.text)
        except Exception as e:
            print("Error conexión:", str(e))

    def cambiar_pagina(self, pagina):
        effect = QGraphicsOpacityEffect(self.contenido.currentWidget())
        self.contenido.currentWidget().setGraphicsEffect(effect)
        self.anim = QPropertyAnimation(effect, b"opacity")
        self.anim.setDuration(200)
        self.anim.setStartValue(1)
        self.anim.setEndValue(0)
        
        def mostrar():
            self.contenido.setCurrentWidget(pagina)
            
            effect2 = QGraphicsOpacityEffect(pagina)
            pagina.setGraphicsEffect(effect2)
            self.anim2 = QPropertyAnimation(effect2, b"opacity")
            self.anim2.setDuration(200)
            self.anim2.setStartValue(0)
            self.anim2.setEndValue(1)
            self.anim2.start()
            
            self.cerrar_menu()
        self.anim.finished.connect(mostrar)
        self.anim.start()
    
    def toggle_menu(self):
        altura = self.submenu.sizeHint().height()
        if self.submenu.maximumHeight() == 0:
            final = altura
        else:
            final = 0
        self.anim_menu = QPropertyAnimation(self.submenu, b"maximumHeight")
        self.anim_menu.setDuration(250)
        self.anim_menu.setStartValue(self.submenu.maximumHeight())
        self.anim_menu.setEndValue(final)
        self.anim_menu.setEasingCurve(QEasingCurve.InOutQuart)
        self.anim_menu.start()

    def cerrar_menu(self):
        if self.submenu.maximumHeight() > 0:
            self.anim_menu = QPropertyAnimation(self.submenu, b"maximumHeight")
            self.anim_menu.setDuration(200)
            self.anim_menu.setStartValue(self.submenu.maximumHeight())
            self.anim_menu.setEndValue(0)
            self.anim_menu.setEasingCurve(QEasingCurve.InOutQuart)
            self.anim_menu.start()

    def subir_archivo(self):
        archivos, _ = QFileDialog.getOpenFileNames(
            self,
            "Seleccionar archivos",
            "",
            "archivos (*.xlsx *.xls *.csv)"
            )
        if archivos:
            for archivo in archivos:
                self.archivos.append(archivo)
                self.lista_archivos.addItem(archivo)
                print("Archivos cargados:", archivo)
            print("Total archivos cargados: ", len(self.archivos))
        else:
            print("Archivo no cargado")

        self.lista_archivos.show()
        self.progress.show()
        self.buscador.hide()
        self.filtro_columna.hide()
        self.tabla.hide()

    def exportar_excel(self):
        ruta, _ = QFileDialog.getSaveFileName(
            self,
            "Guardar archivo",
            "",
            "Excel (*.xlsx)"
        )
        
        if not ruta:
            return
        
        datos = []
        columnas = []
        
        for col in range(self.tabla.columnCount()):
            columnas.append(self.tabla.horizontalHeaderItem(col).text())
        for fila in range(self.tabla.rowCount()):
            fila_datos = []
            for col in range(self.tabla.columnCount()):
                item = self.tabla.item(fila, col)
                fila_datos.append(item.text() if item else "")
            datos.append(fila_datos)
        
        df = pd.DataFrame(datos, columns=columnas)
        df.to_excel(ruta, index=False)
        
        wb = load_workbook(ruta)
        ws = wb.active
        
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="2D89EF", fill_type="solid")
        
        ws.auto_filter.ref = ws.dimensions
        wb.save(ruta)
        
        item = QListWidgetItem()
        widget = DescargaChrome(ruta)
        
        item.setSizeHint(widget.sizeHint())
        
        self.lista_descargas.addItem(item)
        self.lista_descargas.setItemWidget(item, widget)
        
    def mostrar_tabla(self, df):
        self.tabla.setRowCount(len(df))
        self.tabla.setColumnCount(len(df.columns))

        self.tabla.setHorizontalHeaderLabels(df.columns)

        for i in range(len(df)):
            for j in range(len(df.columns)):
                valor = str(df.iloc[i, j])
                self.tabla.setItem(
                    i,
                    j,
                    QTableWidgetItem(valor)
                    )

        self.tabla.resizeColumnsToContents()
        self.tabla.setAlternatingRowColors(True)
        self.filtro_columna.clear()
        self.filtro_columna.addItem("Todas")
        for col in df.columns:
            self.filtro_columna.addItem(col)
        
        self.lista_archivos.hide()
        self.progress.hide()
        self.buscador.show()
        self.filtro_columna.show()
        self.tabla.show()
        
        effect = QGraphicsOpacityEffect(self.tabla)
        self.tabla.setGraphicsEffect(effect)
        
        self.anim = QPropertyAnimation(effect, b"opacity")
        self.anim.setDuration(600)
        self.anim.setStartValue(0)
        self.anim.setEndValue(1)
        self.anim.setEasingCurve(QEasingCurve.InOutQuad)
        self.anim.start()
                
    def buscar_tabla(self):
        texto = self.buscador.text().lower()
        columna = self.filtro_columna.currentIndex() - 1
        for fila in range(self.tabla.rowCount()):
            mostrar = False
            if columna == -1:
                for col in range(self.tabla.columnCount()):
                    item = self.tabla.item(fila, col)
                    if item and texto in item.text().lower():
                        mostrar = True
                        break
            else:
                item = self.tabla.item(fila, columna)
                if item and texto in item.text().lower():
                    mostrar = True
            self.tabla.setRowHidden(fila, not mostrar)
